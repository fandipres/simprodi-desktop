"""
Tools: Hasilkan calon daftar SP_Tahap Awal, SP_Tahap Akhir, dan daftar
mahasiswa non-aktif berkepanjangan untuk semester TERBARU, dari data
mentah mahasiswa aktif, non-aktif, & lulus per semester (format kolom:
Nim, Nama, Program Studi, Ipk, Total Sks Lulus, dst - hasil export
akademik).

Aturan (kebijakan kampus):
- SP Tahap Awal: mahasiswa AKTIF dengan IPK akumulasi < 2.0 pada semester
  itu, DAN semester tempuhnya maksimal 6 (semester 7 ke atas ditangani
  lewat SP Tahap Akhir saja, bukan Tahap Awal lagi). Dicek tiap semester
  secara independen (tidak ada masa berlaku).
- SP Tahap Akhir: mahasiswa aktif ATAU non-aktif yang belum lulus (masih
  ada di salah satu roster, dan TIDAK ada di daftar lulus semester itu)
  begitu semester tempuhnya mencapai 8 atau lebih. Checkpoint berikutnya
  RELATIF terhadap kapan pertama kali kena, bukan daftar semester tetap
  - kena pertama di semester 8 -> checkpoint berikut di 10, 12, dst;
  kena pertama di semester 9 (mis. mahasiswa yang datanya baru muncul
  di term Ganjil) -> checkpoint berikut di 11, 13, dst (masa berlaku
  tiap kena = 2 semester sebelum dicek lagi).
- Data lulus (opsional per semester) dipakai untuk MENGECUALIKAN
  mahasiswa yang sudah lulus semester itu dari SP Tahap Awal maupun
  Akhir - export "aktif" kadang masih memuat mahasiswa yang baru saja
  lulus semester itu juga.
- "SP ke-" dihitung sendiri (bukan input manual): tiap kali seorang
  mahasiswa masuk salah satu daftar SP (Awal atau Akhir) di suatu
  semester, levelnya = level SP terakhir dia + 1, maksimal SP-3.
- Non-aktif berkepanjangan: mahasiswa yang muncul di daftar non-aktif
  pada >= 4 semester BERTURUT-TURUT (dihitung sampai semester non-aktif
  terbaru yang tersedia di antara file yang diberikan).
- DO: dipicu salah satu dari 2 kondisi (mana yang lebih dulu terjadi) -
  (a) mahasiswa yang SUDAH berada di level SP-3 (baik dari riwayat SP
  Tahap Awal maupun SP Tahap Akhir) dan kena checkpoint SP lagi (Tahap
  Awal atau Tahap Akhir, sesuai aturan di atas) - alih-alih naik ke
  "SP-4" yang tidak ada, mahasiswa itu masuk sheet "DO"; atau (b)
  mahasiswa aktif/non-aktif (dan belum lulus) yang semester tempuhnya
  sudah mencapai 14 atau lebih, terlepas dari status SP-nya (backstop).

Opsional: file rekap SP/DO semester-semester SEBELUM data aktif/non-
aktif tersedia (format standar maupun lama - lihat eskalasi_sp.py) bisa
diikutkan lewat parameter recap_paths/--recap, supaya level SP
menyambung dari histori lama (bukan mulai dari nol) dan tiap baris
hasil disertai kolom "Eskalasi" (status naik dari level apa).

Semester tempuh dihitung dari 2 digit awal NIM (tahun masuk) dibandingkan
tahun & term (Ganjil/Genap) nama filenya - bukan dari kolom "Semester"
(kolom itu tidak ada di data mentah aktif/non-aktif). Formula ini sudah
divalidasi cocok 100% dengan kolom "Semester" pada file "Data SP dan DO"
yang sudah ada.

Cara pakai (dari folder ini):
    python generate_sp_do.py --aktif "2025-2026 Genap Aktif.xlsx" ... \\
        --nonaktif "2025-2026 Genap Nonaktif.xlsx" ... \\
        --lulus "2025-2026 Genap Lulus.xlsx" ... --output "hasil.xlsx"
"""

import argparse
import os
import re

import pandas as pd

MAX_SP_LEVEL = 3
AWAL_MAX_SEMESTER = 6
AKHIR_MIN_SEMESTER = 8
AKHIR_EXEMPT_GAP = 2  # semester masa berlaku sebelum checkpoint berikutnya dicek lagi
NONAKTIF_MIN_STREAK = 4
DO_MIN_SEMESTER = 14  # backstop: DO walau belum SP-3, kalau semester tempuh sudah >= ini

_TERM_YEAR_FIRST = re.compile(r"(Ganjil|Genap)\s+(\d{4})-(\d{4})", re.IGNORECASE)
_YEAR_TERM_FIRST = re.compile(r"(\d{4})-(\d{4})\s+(Ganjil|Genap)", re.IGNORECASE)


def parse_term(filename):
    """Kembalikan (year_start, term) dari nama file, atau None kalau tidak cocok pola manapun."""
    name = os.path.basename(filename)
    match = _TERM_YEAR_FIRST.search(name)
    if match:
        return int(match.group(2)), match.group(1).capitalize()
    match = _YEAR_TERM_FIRST.search(name)
    if match:
        return int(match.group(1)), match.group(3).capitalize()
    return None


def scan_folder(folder_path):
    """
    Pindai folder untuk file Excel yang namanya cocok pola semester
    (Ganjil/Genap <tahun>-<tahun>, dua arah urutan). Kembalikan
    (matched, skipped): matched = list path siap dipakai (urut nama),
    skipped = list nama file yang diabaikan (bukan .xls/.xlsx, file
    lock Excel "~$...", atau namanya tidak cocok pola semester).
    """
    matched, skipped = [], []
    with os.scandir(folder_path) as entries:
        for entry in sorted(entries, key=lambda e: e.name):
            if not entry.is_file() or entry.name.startswith("~$"):
                continue
            if not entry.name.lower().endswith((".xls", ".xlsx")):
                skipped.append(entry.name)
                continue
            if parse_term(entry.path) is None:
                skipped.append(entry.name)
                continue
            matched.append(entry.path)
    return matched, skipped


def _semester_label(year_start, term):
    return f"{term} {year_start}/{year_start + 1}"


def _semester_sort_index(year_start, term):
    return year_start * 2 + (0 if term == "Ganjil" else 1)


def _enrollment_year(nim):
    return 2000 + int(nim[:2])


def _semester_number(nim, year_start, term):
    enroll_year = _enrollment_year(nim)
    return (year_start - enroll_year) * 2 + (1 if term == "Ganjil" else 2)


def _read_roster(path):
    """Baca 1 file aktif/non-aktif/lulus mentah, kembalikan DataFrame kolom dinormalisasi."""
    df = pd.read_excel(path)
    # NIM dilewatkan lewat to_numeric+int dulu (bukan str(...) langsung) -
    # kalau kolom "Nim" di file ini kebetulan ada baris kosong, pandas
    # membaca seluruh kolom sebagai float64, dan str() langsung akan
    # menghasilkan NIM salah seperti "231111699.0" (merusak pencocokan
    # NIM lintas file - lihat pola yang sama di eskalasi_sp._get_nim).
    nim_numeric = pd.to_numeric(df["Nim"], errors="coerce")
    df = df[nim_numeric.notna()]
    nim = nim_numeric[nim_numeric.notna()].astype("int64").astype(str)
    out = pd.DataFrame({
        "NIM": nim,
        "Nama": df["Nama"].astype(str).str.strip(),
        "Prodi": df["Program Studi"].astype(str).str.strip(),
        "IPK": pd.to_numeric(df["Ipk"], errors="coerce"),
        "SKS Lulus": pd.to_numeric(df["Total Sks Lulus"], errors="coerce"),
    })
    return out.drop_duplicates(subset="NIM")


def _pair_semesters(aktif_paths, nonaktif_paths, lulus_paths=()):
    """Gabungkan file aktif, non-aktif, & lulus jadi daftar semester terurut kronologis."""
    by_key = {}
    for kind, paths in (("aktif", aktif_paths), ("nonaktif", nonaktif_paths), ("lulus", lulus_paths)):
        for path in paths:
            parsed = parse_term(path)
            if parsed is None:
                raise ValueError(f"Tidak bisa mendeteksi semester dari nama file: {os.path.basename(path)}")
            by_key.setdefault(parsed, {})[kind] = path

    semesters = []
    for (year_start, term), files in sorted(by_key.items(), key=lambda kv: _semester_sort_index(*kv[0])):
        if "aktif" not in files and "nonaktif" not in files:
            continue  # cuma ada file lulus tanpa aktif/nonaktif - tidak ada yang diproses, lewati
        semesters.append({
            "year_start": year_start,
            "term": term,
            "label": _semester_label(year_start, term),
            "aktif_path": files.get("aktif"),
            "nonaktif_path": files.get("nonaktif"),
            "lulus_path": files.get("lulus"),
        })
    return semesters


def _level_label(level):
    return "DO" if level > MAX_SP_LEVEL else f"SP-{level}"


def _eskalasi_note(prev_level, new_level):
    new_label = _level_label(new_level)
    if prev_level == 0:
        return f"Baru ({new_label})"
    if new_level > prev_level:
        return f"{_level_label(prev_level)} -> {new_label}"
    return f"Tetap {new_label}"


def _is_escalation(prev_level, new_level):
    """Eskalasi sungguhan: sudah pernah kena sebelumnya (prev_level > 0) DAN levelnya naik."""
    return prev_level > 0 and new_level > prev_level


def _sp_row(row, nim, semester_number, level, status, prev_level=None):
    result = {
        "NIM": nim,
        "Nama": row["Nama"],
        "Prodi": row["Prodi"],
        "Semester": semester_number,
        "IPK": round(row["IPK"], 4) if pd.notna(row["IPK"]) else "",
        "SKS Lulus": int(row["SKS Lulus"]) if pd.notna(row["SKS Lulus"]) else "",
        "SP ke-": level,
        "Status": status,
    }
    if prev_level is not None:
        result["Eskalasi"] = _eskalasi_note(prev_level, level)
    return result


def _do_row(row, nim, semester_number, sumber, status, prev_level=None):
    if sumber == "semester":
        keterangan = f"Sudah memasuki semester {semester_number} tanpa lulus"
    else:
        keterangan = f"Sudah SP-3 (dari SP {sumber}), belum lulus di semester {semester_number}"
    result = {
        "NIM": nim,
        "Nama": row["Nama"],
        "Prodi": row["Prodi"],
        "Semester": semester_number,
        "IPK": round(row["IPK"], 4) if pd.notna(row["IPK"]) else "",
        "SKS Lulus": int(row["SKS Lulus"]) if pd.notna(row["SKS Lulus"]) else "",
        "Keterangan": keterangan,
        "Status": status,
    }
    if prev_level is not None:
        result["Eskalasi"] = _eskalasi_note(prev_level, MAX_SP_LEVEL + 1)
    return result


def generate(aktif_paths, nonaktif_paths, lulus_paths=(), recap_paths=()):
    """
    Kembalikan dict hasil untuk semester TERBARU:
        label, semester_diproses, sp_tahap_awal, sp_tahap_akhir, do,
        nonaktif_berkepanjangan, has_eskalasi (list of dict, siap
        dipakai df/ekspor). Tiap baris SP/DO disertai kolom "Status"
        (Aktif/Non-Aktif pada semester itu).

    Kalau recap_paths diisi, level SP disambung dari riwayat file rekap
    tersebut (eskalasi_sp.compute_levels) dan tiap baris SP/DO hasil
    disertai kolom tambahan "Eskalasi" (status naik dari level apa
    sebelumnya) di ujung kanan.
    """
    semesters = _pair_semesters(aktif_paths, nonaktif_paths, lulus_paths)
    if not semesters:
        raise ValueError("Tidak ada file aktif/non-aktif yang valid.")

    has_eskalasi = bool(recap_paths)
    recap_labels = []
    if has_eskalasi:
        from core import eskalasi_sp

        running_level, _recap_info, recap_labels, _last_sp_key = eskalasi_sp.compute_levels(list(recap_paths))
        # Sambungkan masa berlaku Tahap Akhir dari rekap juga - supaya
        # mahasiswa yang kena Tahap Akhir di semester yang datanya
        # berasal dari rekap tidak dianggap "belum pernah kena" begitu
        # masuk ke semester yang diproses live (lihat catatan di
        # eskalasi_sp.compute_levels).
        akhir_last_hit = {
            nim: key[0] * 2 + key[1] for nim, key in _last_sp_key.items() if key[0] != 9999
        }
        recap_counts = eskalasi_sp.semester_counts(list(recap_paths))
    else:
        running_level = {}    # nim -> level SP terakhir (0 kalau belum pernah)
        akhir_last_hit = {}   # nim -> sort index semester terakhir kena Tahap Akhir
        recap_counts = {}     # label -> {"sp": ..., "do": ...} (jumlah unik NIM di file rekap)

    nonaktif_streak = {}  # nim -> jumlah semester non-aktif berturut-turut sampai sekarang
    last_info = {}        # nim -> {nama, prodi}
    semester_summary = [] # rekap jumlah aktif/nonaktif/lulus/sp/do per semester (buat sheet Infografis)

    result = {}

    for sem in semesters:
        sort_index = _semester_sort_index(sem["year_start"], sem["term"])

        empty = pd.DataFrame(columns=["NIM", "Nama", "Prodi", "IPK", "SKS Lulus"])
        aktif_df = _read_roster(sem["aktif_path"]) if sem["aktif_path"] else empty
        nonaktif_df = _read_roster(sem["nonaktif_path"]) if sem["nonaktif_path"] else empty
        lulus_df = _read_roster(sem["lulus_path"]) if sem["lulus_path"] else empty
        lulus_nims = set(lulus_df["NIM"])

        non_empty = [df for df in (aktif_df, nonaktif_df) if not df.empty]
        combined_df = pd.concat(non_empty).drop_duplicates(subset="NIM") if non_empty else empty

        status_map = {nim: "Non-Aktif" for nim in nonaktif_df["NIM"]}
        status_map.update({nim: "Aktif" for nim in aktif_df["NIM"]})

        for _, row in combined_df.iterrows():
            last_info[row["NIM"]] = {"nama": row["Nama"], "prodi": row["Prodi"]}

        nonaktif_nims = set(nonaktif_df["NIM"])
        for nim in nonaktif_nims:
            nonaktif_streak[nim] = nonaktif_streak.get(nim, 0) + 1
        for nim in list(nonaktif_streak):
            if nim not in nonaktif_nims:
                nonaktif_streak[nim] = 0

        summary_entry = {
            "label": sem["label"],
            "aktif": len(aktif_df),
            "nonaktif": len(nonaktif_df),
            "lulus": len(lulus_df),
            "sp": 0,
            "do": 0,
        }
        semester_summary.append(summary_entry)

        if sem["label"] in recap_labels:
            # Semester ini sudah tercakup file rekap (folder riwayat
            # non-aktif/lulus kerap tumpang tindih periode dengan rekap
            # SP/DO - ini malah kasus UMUM, bukan langka) - level sudah
            # otoritatif dari rekap, jangan dihitung ulang dari aturan
            # mentah supaya tidak menimpa data resmi. Streak non-aktif di
            # atas tetap jalan karena rekap tidak punya info itu. Jumlah
            # SP/DO buat summary_entry diambil dari file rekap juga
            # (recap_counts), bukan dihitung ulang.
            counts = recap_counts.get(sem["label"], {"sp": 0, "do": 0})
            summary_entry["sp"] = counts["sp"]
            summary_entry["do"] = counts["do"]
            continue

        do_rows, do_eskalasi, do_nims = [], [], set()

        def _maybe_do(row, nim, semester_number, sumber, prev_level):
            """Kalau prev_level sudah SP-3, ini jadi DO (bukan naik ke SP-4 yang tidak ada)."""
            if prev_level < MAX_SP_LEVEL:
                return False
            status = status_map.get(nim, "Aktif")
            do_row = _do_row(row, nim, semester_number, sumber, status, prev_level if has_eskalasi else None)
            do_rows.append(do_row)
            do_nims.add(nim)
            running_level[nim] = MAX_SP_LEVEL + 1
            if has_eskalasi and _is_escalation(prev_level, MAX_SP_LEVEL + 1):
                do_eskalasi.append(do_row)
            return True

        awal_rows, awal_eskalasi = [], []
        for _, row in aktif_df.iterrows():
            nim = row["NIM"]
            if nim in lulus_nims:
                continue  # sudah lulus semester ini, bukan kandidat SP
            semester_number = _semester_number(nim, sem["year_start"], sem["term"])
            if semester_number > AWAL_MAX_SEMESTER:
                continue  # semester 7 ke atas ditangani lewat Tahap Akhir, bukan Tahap Awal
            if pd.notna(row["IPK"]) and row["IPK"] < 2.0:
                prev_level = running_level.get(nim, 0)
                if _maybe_do(row, nim, semester_number, "Tahap Awal", prev_level):
                    continue
                level = prev_level + 1
                running_level[nim] = level
                sp_row = _sp_row(row, nim, semester_number, level, "Aktif", prev_level if has_eskalasi else None)
                awal_rows.append(sp_row)
                if has_eskalasi and _is_escalation(prev_level, level):
                    awal_eskalasi.append(sp_row)

        akhir_rows, akhir_eskalasi = [], []
        for _, row in combined_df.iterrows():
            nim = row["NIM"]
            if nim in lulus_nims or nim in do_nims:
                continue  # sudah lulus, atau sudah kena DO lewat Tahap Awal semester ini
            semester_number = _semester_number(nim, sem["year_start"], sem["term"])
            if semester_number < AKHIR_MIN_SEMESTER:
                continue
            last_hit = akhir_last_hit.get(nim)
            if last_hit is not None and sort_index - last_hit < AKHIR_EXEMPT_GAP:
                continue  # masih dalam masa berlaku SP Tahap Akhir sebelumnya
            akhir_last_hit[nim] = sort_index
            prev_level = running_level.get(nim, 0)
            if _maybe_do(row, nim, semester_number, "Tahap Akhir", prev_level):
                continue
            level = prev_level + 1
            running_level[nim] = level
            status = status_map.get(nim, "Aktif")
            sp_row = _sp_row(row, nim, semester_number, level, status, prev_level if has_eskalasi else None)
            akhir_rows.append(sp_row)
            if has_eskalasi and _is_escalation(prev_level, level):
                akhir_eskalasi.append(sp_row)

        # DO backstop: semester tempuh sudah >= 14, terlepas dari status SP,
        # asal belum ke-DO-kan lewat aturan SP-3 di atas semester ini.
        for _, row in combined_df.iterrows():
            nim = row["NIM"]
            if nim in lulus_nims or nim in do_nims:
                continue
            semester_number = _semester_number(nim, sem["year_start"], sem["term"])
            if semester_number >= DO_MIN_SEMESTER:
                prev_level = running_level.get(nim, 0)
                status = status_map.get(nim, "Aktif")
                do_row = _do_row(row, nim, semester_number, "semester", status, prev_level if has_eskalasi else None)
                do_rows.append(do_row)
                do_nims.add(nim)
                running_level[nim] = MAX_SP_LEVEL + 1
                if has_eskalasi and _is_escalation(prev_level, MAX_SP_LEVEL + 1):
                    do_eskalasi.append(do_row)

        summary_entry["sp"] = len(awal_rows) + len(akhir_rows)
        summary_entry["do"] = len(do_rows)

        result = {
            "label": sem["label"],
            "sp_tahap_awal": awal_rows,
            "sp_tahap_awal_eskalasi": awal_eskalasi,
            "sp_tahap_akhir": akhir_rows,
            "sp_tahap_akhir_eskalasi": akhir_eskalasi,
            "do": do_rows,
            "do_eskalasi": do_eskalasi,
        }

    nonaktif_rows = []
    for nim, streak in nonaktif_streak.items():
        if streak >= NONAKTIF_MIN_STREAK:
            info = last_info.get(nim, {"nama": "", "prodi": ""})
            nonaktif_rows.append({
                "NIM": nim,
                "Nama": info["nama"],
                "Prodi": info["prodi"],
                "Semester Non-Aktif Berturut-turut": streak,
            })
    nonaktif_rows.sort(key=lambda r: r["NIM"])

    result["nonaktif_berkepanjangan"] = nonaktif_rows
    result["semester_diproses"] = recap_labels + [
        s["label"] for s in semesters if s["label"] not in recap_labels
    ]
    result["has_eskalasi"] = has_eskalasi
    result["semester_summary"] = semester_summary
    return result


# Lebar kolom persis mengikuti Template.xlsx yang diberikan (per sheet,
# karena tiap sheet template punya lebar kolom sedikit berbeda). Kolom
# "Status" (Aktif/Non-Aktif) dan "Eskalasi" (kalau mode B) ditambahkan
# di ujung, di luar kolom asli template.
_TEMPLATE_COLUMN_WIDTHS = {
    "SP_Tahap Awal": [10.0, 34.44, 19.44, 13.33, 12.0, 13.33, 10.66],
    "SP_Tahap Akhir": [10.0, 29.33, 19.44, 13.33, 12.0, 13.33, 10.66],
    "DO": [10.0, 23.66, 19.44, 8.66, 12.0, 13.0, 37.33],
}
_STATUS_COLUMN_WIDTH = 12.0
_ESKALASI_COLUMN_WIDTH = 22.0
_INSIGHT_COLUMN_WIDTH = 70.0


def _build_insight(category, labels, values):
    """
    Ringkasan tren sederhana berbasis aturan (bukan model AI sungguhan -
    aplikasi ini jalan offline, tidak ada panggilan ke layanan AI) untuk 1
    baris kategori di sheet Infografis, supaya pembaca cepat menangkap arah
    perubahan tanpa perlu baca grafiknya baris per baris. labels/values
    HARUS berurutan kronologis (lama -> baru), beda dari urutan tampilan
    tabel yang baru -> lama.
    """
    if len(values) < 2:
        return "Data cuma 1 semester, belum bisa dilihat trennya."

    latest, prev = values[-1], values[-2]
    latest_label, prev_label = labels[-1], labels[-2]
    delta = latest - prev
    if delta > 0:
        arah = f"naik {delta}"
    elif delta < 0:
        arah = f"turun {abs(delta)}"
    else:
        arah = "sama"

    max_idx = max(range(len(values)), key=lambda i: values[i])
    min_idx = min(range(len(values)), key=lambda i: values[i])

    mid = len(values) // 2
    first_half_avg = sum(values[:mid]) / mid
    second_half_avg = sum(values[mid:]) / (len(values) - mid)
    if second_half_avg > first_half_avg * 1.1:
        tren = "cenderung meningkat"
    elif second_half_avg < first_half_avg * 0.9:
        tren = "cenderung menurun"
    else:
        tren = "relatif stabil"

    return (
        f"{category} semester {latest_label}: {latest} ({arah} dari {prev_label}). "
        f"Tertinggi di {labels[max_idx]} ({values[max_idx]}), terendah di "
        f"{labels[min_idx]} ({values[min_idx]}). Tren keseluruhan {tren}."
    )


def export_excel(result, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    has_eskalasi = result.get("has_eskalasi", False)
    sp_columns = ["NIM", "Nama", "Prodi", "Semester", "IPK", "SKS Lulus", "SP ke-", "Status"]
    do_columns = ["NIM", "Nama", "Prodi", "Semester", "IPK", "SKS Lulus", "Keterangan", "Status"]
    nonaktif_columns = ["NIM", "Nama", "Prodi", "Semester Non-Aktif Berturut-turut"]
    if has_eskalasi:
        sp_columns = sp_columns + ["Eskalasi"]
        do_columns = do_columns + ["Eskalasi"]

    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(style="thin")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9D9D9")  # sesuai Template.xlsx
    escalation_fill = PatternFill("solid", fgColor="FFF2CC")
    center = Alignment(horizontal="center")

    def write_sheet(name, rows, columns, highlight_col=None, column_widths=None):
        ws = wb.create_sheet(name)
        df = pd.DataFrame(rows, columns=columns)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = cell_border
            cell.alignment = center

        highlight_idx = columns.index(highlight_col) + 1 if highlight_col else None
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = cell_border
                cell.alignment = center
                if highlight_idx is not None and col_idx == highlight_idx:
                    cell.fill = escalation_fill

        for col_idx, col_name in enumerate(columns, start=1):
            if column_widths is not None:
                width = column_widths[col_idx - 1]
            else:
                width = max(12, min(40, len(col_name) + 4))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    def write_infografis_sheet(semester_summary):
        from openpyxl.chart import BarChart, Reference

        ws = wb.create_sheet("Infografis")

        if not semester_summary:
            ws.cell(row=1, column=1, value="Tidak ada data semester untuk direkap.")
            return

        # "Aktif" sengaja tidak direkap di sini - tool ini cuma diberi file
        # aktif untuk semester yang sedang diproses (tidak ada folder
        # riwayat aktif seperti Non-Aktif/Lulus), jadi angkanya akan 0 di
        # hampir semua semester dan menyesatkan kalau ditampilkan/dianalisis.
        rows = [("Non-Aktif", "nonaktif"), ("Lulus", "lulus"), ("SP", "sp"), ("DO", "do")]

        # semester_summary datang dari generate() berurutan kronologis
        # (lama -> baru, dibutuhkan buat hitung tren di _build_insight).
        # Tabel & grafik ditampilkan sebaliknya, semester TERBARU dulu.
        chronological = semester_summary
        display = list(reversed(semester_summary))

        n_rows = len(rows) + 1
        n_cols = len(display) + 1
        insight_col = n_cols + 1

        ws.cell(row=1, column=1, value="Kategori")
        for col, entry in enumerate(display, start=2):
            ws.cell(row=1, column=col, value=entry["label"])
        ws.cell(row=1, column=insight_col, value="Insight (Ringkasan Otomatis)")

        for row_idx, (label, key) in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=label)
            for col, entry in enumerate(display, start=2):
                ws.cell(row=row_idx, column=col, value=entry[key])
            values = [entry[key] for entry in chronological]
            labels = [entry["label"] for entry in chronological]
            ws.cell(row=row_idx, column=insight_col, value=_build_insight(label, labels, values))

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = cell_border
            cell.alignment = center
        for row_idx in range(2, n_rows + 1):
            for col_idx in range(1, n_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = cell_border
                cell.alignment = center
                if col_idx == 1:
                    cell.font = Font(bold=True)
            insight_cell = ws.cell(row=row_idx, column=insight_col)
            insight_cell.border = cell_border
            insight_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        ws.column_dimensions["A"].width = 14
        for col_idx in range(2, n_cols + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16
        ws.column_dimensions[ws.cell(row=1, column=insight_col).column_letter].width = _INSIGHT_COLUMN_WIDTH
        for row_idx in range(2, n_rows + 1):
            ws.row_dimensions[row_idx].height = 45

        # Kolom SP/DO bisa 0 untuk semester yang datanya berasal dari file
        # rekap (bukan dihitung ulang dari data mentah) - lihat catatan di
        # generate() soal summary_entry pada semester yang ada di recap_labels.
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = "Rekap Mahasiswa per Semester"
        chart.y_axis.title = "Jumlah Mahasiswa"
        chart.x_axis.title = "Semester"
        chart.style = 10

        data = Reference(ws, min_col=1, max_col=n_cols, min_row=2, max_row=n_rows)
        chart.add_data(data, titles_from_data=True, from_rows=True)
        cats = Reference(ws, min_col=2, max_col=n_cols, min_row=1, max_row=1)
        chart.set_categories(cats)

        chart.width = max(20, n_cols * 3)
        chart.height = 12

        ws.add_chart(chart, f"A{n_rows + 3}")

    extra_widths = [_STATUS_COLUMN_WIDTH] + ([_ESKALASI_COLUMN_WIDTH] if has_eskalasi else [])
    highlight_col = "Eskalasi" if has_eskalasi else None

    write_sheet("SP_Tahap Awal", result["sp_tahap_awal"], sp_columns, highlight_col=highlight_col,
                column_widths=_TEMPLATE_COLUMN_WIDTHS["SP_Tahap Awal"] + extra_widths)
    write_sheet("SP_Tahap Akhir", result["sp_tahap_akhir"], sp_columns, highlight_col=highlight_col,
                column_widths=_TEMPLATE_COLUMN_WIDTHS["SP_Tahap Akhir"] + extra_widths)
    write_sheet("DO", result["do"], do_columns, highlight_col=highlight_col,
                column_widths=_TEMPLATE_COLUMN_WIDTHS["DO"] + extra_widths)

    write_sheet("Rekap Non-Aktif", result["nonaktif_berkepanjangan"], nonaktif_columns)

    write_infografis_sheet(result.get("semester_summary", []))

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--aktif", nargs="+", default=[], help="Path file mahasiswa aktif per semester")
    parser.add_argument("--nonaktif", nargs="+", default=[], help="Path file mahasiswa non-aktif per semester")
    parser.add_argument("--lulus", nargs="+", default=[], help="Path file mahasiswa lulus per semester (opsional)")
    parser.add_argument("--recap", nargs="+", default=[], help="Path file rekap SP/DO semester sebelumnya (opsional, untuk menyambung histori & kolom Eskalasi)")
    parser.add_argument("--output", required=True, help="Path file Excel hasil")
    args = parser.parse_args()

    result = generate(args.aktif, args.nonaktif, args.lulus, args.recap)
    export_excel(result, args.output)
    print(f"Semester diproses : {', '.join(result['semester_diproses'])}")
    print(f"Semester hasil     : {result['label']}")
    print(f"SP Tahap Awal      : {len(result['sp_tahap_awal'])}")
    print(f"SP Tahap Akhir     : {len(result['sp_tahap_akhir'])}")
    print(f"DO                 : {len(result['do'])}")
    print(f"Rekap Non-Aktif    : {len(result['nonaktif_berkepanjangan'])}")
    print(f"Hasil : {args.output}")


if __name__ == "__main__":
    main()
