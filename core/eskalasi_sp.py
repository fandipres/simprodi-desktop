"""
Tools: Deteksi eskalasi Surat Peringatan (SP) mahasiswa lintas semester.

Baca beberapa file rekap SP per semester sebagai semacam database
berurut. Tiga variasi sheet dikenali sekaligus (boleh dicampur di
beberapa file berbeda):

- Format standar (sejak Ganjil 2024/2025): sheet "SP_Tahap Awal",
  "SP_Tahap Akhir", "DO" - semuanya dianggap milik semester sesuai nama
  filenya ("Ganjil/Genap <tahun>-<tahun>" atau "<tahun>-<tahun> Ganjil/
  Genap").
- Format histori yang sudah diseragamkan: sheet "SP" (daftar SP
  gabungan, tanpa pemisahan Awal/Akhir) dan "DO" - juga dianggap milik
  semester sesuai nama file.
- Format lama asli (kalau masih ada): sheet "List Peringatan <Ganjil/
  Genap> <yy><yy>" (daftar SP gabungan) dan "DO SP<n> <Ganjil/Genap>
  <yy><yy>" (daftar DO). Semester tiap sheet format ini diambil dari
  NAMA SHEET itu sendiri (bukan nama file) - sheet DO kadang disimpan
  sebagai arsip di file semester lain, jadi tidak bisa diasumsikan sama
  dengan semester filenya.

Level SP TIDAK dibaca mentah-mentah dari kolom "SP ke-"/"SP" (kolom itu
bisa salah input, mis. ditemukan "SP-4" yang seharusnya tidak ada),
melainkan dihitung sendiri dari kemunculan NIM di tiap semester secara
berurutan: kalau semester ini seorang mahasiswa muncul di daftar SP,
levelnya = level semester sebelumnya + 1 (mahasiswa baru mulai dari
SP-1). Level dibatasi maksimal SP-3 - setelah itu status berikutnya
seharusnya DO/mengundurkan diri, bukan SP-4.

Mahasiswa yang levelnya di semester terakhir lebih tinggi dari level
tertinggi di semester-semester sebelumnya ditandai sebagai "eskalasi".
DO dianggap level paling tinggi (di atas SP berapa pun).
"""

import argparse
import os
import re

import pandas as pd

MAX_SP_LEVEL = 3
DO_RANK = MAX_SP_LEVEL + 1

_TERM_YEAR_FIRST = re.compile(r"(Ganjil|Genap)\s+(\d{4})-(\d{4})", re.IGNORECASE)
_YEAR_TERM_FIRST = re.compile(r"(\d{4})-(\d{4})\s+(Ganjil|Genap)", re.IGNORECASE)
_LEGACY_LIST = re.compile(r"List\s+Peringatan\s+(Ganjil|Genap)\s+(\d{2})(\d{2})", re.IGNORECASE)
_LEGACY_DO = re.compile(r"DO\s+SP\d+\s+(Ganjil|Genap)\s+(\d{2})(\d{2})", re.IGNORECASE)

SP_SHEETS = ("SP_Tahap Awal", "SP_Tahap Akhir", "SP")
DO_SHEET = "DO"


def _semester_from_term_year(term, year_start):
    term = term.capitalize()
    sort_key = (year_start, 0 if term == "Ganjil" else 1)
    return sort_key, f"{term} {year_start}/{year_start + 1}"


def _semester_label(path, fallback_order):
    name = os.path.basename(path)
    match = _TERM_YEAR_FIRST.search(name)
    if match:
        return _semester_from_term_year(match.group(1), int(match.group(2)))
    match = _YEAR_TERM_FIRST.search(name)
    if match:
        return _semester_from_term_year(match.group(3), int(match.group(1)))
    return (9999, fallback_order), os.path.splitext(name)[0]


def _parse_legacy_sheet(sheet_name):
    """Kembalikan (sort_key, label, kind) dari nama sheet format lama, atau None."""
    name = sheet_name.strip()
    match = _LEGACY_LIST.match(name)
    if match:
        sort_key, label = _semester_from_term_year(match.group(1), 2000 + int(match.group(2)))
        return sort_key, label, "sp"
    match = _LEGACY_DO.match(name)
    if match:
        sort_key, label = _semester_from_term_year(match.group(1), 2000 + int(match.group(2)))
        return sort_key, label, "do"
    return None


def _find_sheet(xls, wanted_name):
    wanted = wanted_name.strip().lower()
    for sheet in xls.sheet_names:
        if sheet.strip().lower() == wanted:
            return sheet
    return None


def _get_field(row, *column_names):
    for name in column_names:
        if name in row and pd.notna(row[name]):
            text = str(row[name]).strip()
            if text:
                return text
    return ""


def _get_nim(row):
    for name in ("NIM", "Nim"):
        if name in row and pd.notna(row[name]):
            return str(int(row[name]))
    return None


def _collect_file_contributions(path, fallback_order):
    """
    Kembalikan list of (sort_key, label, kind, dataframe) dari 1 file -
    kind "sp" atau "do". Format standar diprioritaskan; kalau tidak ada,
    coba format lama (semester per sheet, lihat docstring modul).
    """
    xls = pd.ExcelFile(path)
    contributions = []

    has_standard = any(_find_sheet(xls, s) for s in SP_SHEETS) or _find_sheet(xls, DO_SHEET)
    if has_standard:
        own_key, own_label = _semester_label(path, fallback_order)
        for sheet_name in SP_SHEETS:
            actual = _find_sheet(xls, sheet_name)
            if actual is not None:
                contributions.append((own_key, own_label, "sp", pd.read_excel(xls, sheet_name=actual)))
        actual_do = _find_sheet(xls, DO_SHEET)
        if actual_do is not None:
            contributions.append((own_key, own_label, "do", pd.read_excel(xls, sheet_name=actual_do)))
        return contributions

    for sheet_name in xls.sheet_names:
        parsed = _parse_legacy_sheet(sheet_name)
        if parsed is None:
            continue
        sort_key, label, kind = parsed
        contributions.append((sort_key, label, kind, pd.read_excel(xls, sheet_name=sheet_name)))

    return contributions


def _blank_presence():
    return {"nama": "", "prodi": "", "dosen_wali": "", "on_sp": False, "on_do": False}


def _group_semester_data(file_paths):
    """Kumpulkan kontribusi semua file jadi dict sort_key -> {label, sp_rows, do_rows}."""
    semester_data = {}
    for i, path in enumerate(file_paths):
        for sort_key, label, kind, df in _collect_file_contributions(path, i):
            bucket = semester_data.setdefault(sort_key, {"label": label, "sp_rows": [], "do_rows": []})
            bucket[f"{kind}_rows"].append(df)
    return semester_data


def _presence_for_semester(bucket):
    """Bangun dict nim -> presence (nama/prodi/dosen_wali/on_sp/on_do) dari 1 bucket semester."""
    presence = {}
    for kind, rows_key in (("sp", "sp_rows"), ("do", "do_rows")):
        for df in bucket[rows_key]:
            for _, row in df.iterrows():
                nim = _get_nim(row)
                if nim is None:
                    continue
                p = presence.setdefault(nim, _blank_presence())
                p["nama"] = _get_field(row, "Nama") or p["nama"]
                p["prodi"] = _get_field(row, "Prodi", "Program Studi") or p["prodi"]
                p["dosen_wali"] = _get_field(row, "Dosen Wali", "Penasehat Akademik") or p["dosen_wali"]
                p["on_" + kind] = True
    return presence


def compute_levels(file_paths):
    """
    Baca file rekap SP/DO (format standar/lama) dan kembalikan
    (running_level, student_info, semester_labels, last_sp_key) hasil
    akhir setelah diproses berurutan - dipakai untuk menyambung riwayat
    SP sebelum data aktif/non-aktif mentah tersedia (lihat
    generate_sp_do.py).

    running_level: {nim: level SP terakhir (0..MAX_SP_LEVEL) atau DO_RANK}
    student_info: {nim: {nama, prodi, dosen_wali}}
    semester_labels: label semester terurut kronologis yang terbaca.
    last_sp_key: {nim: sort_key semester terakhir muncul di SP/DO} -
        dipakai generate_sp_do.py untuk menyambung masa berlaku SP
        Tahap Akhir (supaya mahasiswa yang kena Tahap Akhir di semester
        yang datanya berasal dari rekap tidak dianggap "belum pernah
        kena" begitu masuk ke semester yang diproses live).
    """
    semester_data = _group_semester_data(file_paths)
    ordered_keys = sorted(semester_data.keys())

    running_level = {}
    student_info = {}
    last_sp_key = {}

    for key in ordered_keys:
        presence = _presence_for_semester(semester_data[key])
        for nim, p in presence.items():
            info = student_info.setdefault(nim, {"nama": "", "prodi": "", "dosen_wali": ""})
            for field in ("nama", "prodi", "dosen_wali"):
                if p[field]:
                    info[field] = p[field]
            if p["on_do"]:
                running_level[nim] = DO_RANK
                last_sp_key[nim] = key
            elif p["on_sp"]:
                running_level[nim] = min(running_level.get(nim, 0) + 1, MAX_SP_LEVEL)
                last_sp_key[nim] = key

    semester_labels = [semester_data[key]["label"] for key in ordered_keys]
    return running_level, student_info, semester_labels, last_sp_key


def semester_counts(file_paths):
    """
    Kembalikan {label: {"sp": jumlah NIM unik di sheet SP, "do": jumlah
    NIM unik di sheet DO}} per semester dari file rekap - dipakai buat
    rekap jumlah SP/DO per semester (sheet "Infografis" di
    generate_sp_do.py), BUKAN buat menghitung level (lihat
    compute_levels/find_escalations di atas).
    """
    semester_data = _group_semester_data(file_paths)
    counts = {}
    for bucket in semester_data.values():
        sp_nims = {nim for df in bucket["sp_rows"] for nim in df.apply(_get_nim, axis=1) if nim is not None}
        do_nims = {nim for df in bucket["do_rows"] for nim in df.apply(_get_nim, axis=1) if nim is not None}
        counts[bucket["label"]] = {"sp": len(sp_nims), "do": len(do_nims)}
    return counts


def find_escalations(file_paths):
    """
    Kembalikan (rows, semester_labels). Semester (dari file/sheet)
    diurutkan kronologis; semester PALING BARU dianggap "semester ini".
    rows berisi satu baris per mahasiswa yang ada di daftar SP/DO
    semester ini DAN levelnya (hasil hitung sendiri) lebih tinggi dari
    level tertinggi di semester-semester sebelumnya. semester_labels:
    label semester terurut kronologis.
    """
    semester_data = _group_semester_data(file_paths)

    if len(semester_data) < 2:
        raise ValueError("Perlu minimal 2 semester (dari sheet yang dikenali) untuk mendeteksi eskalasi.")

    ordered_keys = sorted(semester_data.keys())
    semester_labels = [semester_data[key]["label"] for key in ordered_keys]

    student_info = {}  # nim -> {nama, prodi, dosen_wali}
    history = {}  # nim -> {semester_label: (level_rank, level_label)}
    running_level = {}  # nim -> level SP terakhir yang sudah diketahui (0 kalau belum pernah)

    for key in ordered_keys:
        label = semester_data[key]["label"]
        presence = _presence_for_semester(semester_data[key])

        for nim, p in presence.items():
            info = student_info.setdefault(nim, {"nama": "", "prodi": "", "dosen_wali": ""})
            for field in ("nama", "prodi", "dosen_wali"):
                if p[field]:
                    info[field] = p[field]

            if p["on_do"]:
                level_rank, level_label = DO_RANK, "DO"
            elif p["on_sp"]:
                level_rank = min(running_level.get(nim, 0) + 1, MAX_SP_LEVEL)
                level_label = f"SP-{level_rank}"
            else:
                continue

            running_level[nim] = level_rank
            history.setdefault(nim, {})[label] = (level_rank, level_label)

    latest_label = semester_labels[-1]
    earlier_labels = semester_labels[:-1]

    rows = []
    for nim, sem_history in history.items():
        if latest_label not in sem_history:
            continue

        previous_entries = {label: sem_history[label] for label in earlier_labels if label in sem_history}
        if not previous_entries:
            continue  # baru pertama kali kena SP/DO, bukan lanjutan

        prev_max_level = max(level_rank for level_rank, _ in previous_entries.values())
        latest_level, latest_label_text = sem_history[latest_label]
        if latest_level <= prev_max_level:
            continue

        last_seen_label = next(label for label in reversed(earlier_labels) if label in previous_entries)

        info = student_info[nim]
        row = {"NIM": nim, "Nama": info["nama"], "Prodi": info["prodi"], "Dosen Wali": info["dosen_wali"]}
        for label in earlier_labels:
            row[label] = previous_entries[label][1] if label in previous_entries else "-"
        row[latest_label] = latest_label_text
        row["Eskalasi"] = f"{previous_entries[last_seen_label][1]} -> {latest_label_text}"
        rows.append(row)

    rows.sort(key=lambda r: r["NIM"])
    return rows, semester_labels


def export_excel(rows, semester_labels, output_path):
    columns = ["NIM", "Nama", "Prodi", "Dosen Wali", *semester_labels, "Eskalasi"]
    df = pd.DataFrame(rows, columns=columns)
    df.to_excel(output_path, index=False, sheet_name="Eskalasi SP")

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    wb = load_workbook(output_path)
    ws = wb.active
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    escalation_fill = PatternFill("solid", fgColor="FFF2CC")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    escalation_col = columns.index("Eskalasi") + 1
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=escalation_col).fill = escalation_fill

    for col_idx, col_name in enumerate(columns, start=1):
        width = max(12, min(40, len(col_name) + 4))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("files", nargs="+", help="Path file rekap SP per semester (urutan bebas, diurutkan otomatis)")
    parser.add_argument("--output", required=True, help="Path file Excel hasil")
    args = parser.parse_args()

    rows, semester_labels = find_escalations(args.files)
    export_excel(rows, semester_labels, args.output)
    print(f"Semester dibaca : {', '.join(semester_labels)}")
    print(f"Mahasiswa eskalasi : {len(rows)}")
    print(f"Hasil : {args.output}")


if __name__ == "__main__":
    main()
